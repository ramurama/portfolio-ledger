"""Excel renderer using openpyxl.

The renderer supports two modes through a single entry-point:

    * Single-section reports (e.g. Current Holdings) - pass one
      `ExcelSection` and the workbook gets one sheet.
    * Multi-section reports (e.g. per-account Tax Lots) - pass several
      `ExcelSection` objects and each lands on its own sheet.

Each sheet has the same quality-of-life touches:

    * Bold header row with a coloured fill.
    * Frozen first row so headers stay visible while scrolling.
    * Column widths approximated from the longest cell content.
    * Right-aligned numeric columns.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.logging import get_logger

logger = get_logger(__name__)


# Styling constants kept at module level so they can be tweaked once
# and reflected across every Excel report.
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_NUMERIC_ALIGNMENT = Alignment(horizontal="right")


@dataclass(frozen=True)
class ExcelSection:
    """One worksheet's worth of content."""

    sheet_name: str
    headers: list[str]
    body: list[list[str]]


def write_excel(
    output_path: Path,
    sections: Sequence[ExcelSection],
) -> Path:
    """Write a workbook with one sheet per `ExcelSection` to `output_path`.

    `sections` must contain at least one entry. Empty sections still
    produce a sheet with the headers - this matters for per-account
    Tax Lots reports where an account may legitimately have zero trades.
    """

    if not sections:
        raise ValueError("write_excel requires at least one ExcelSection")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Workbook() always starts with one default sheet; we replace its
    # title / content with the first section to avoid an empty sheet
    # being left behind.
    for idx, section in enumerate(sections):
        if idx == 0:
            ws = wb.active
            if ws is None:  # pragma: no cover - defensive
                ws = wb.create_sheet()
        else:
            ws = wb.create_sheet()

        # Excel caps sheet names at 31 characters and forbids a few
        # punctuation chars. Strip those defensively.
        ws.title = _sanitize_sheet_name(section.sheet_name)

        _write_headers(ws, section.headers)
        _write_body(ws, section.headers, section.body)
        _apply_column_widths(ws, section.headers, section.body)

        # Freeze the header row - "A2" means "scroll body, keep row 1".
        ws.freeze_panes = "A2"

    wb.save(output_path)
    total_rows = sum(len(s.body) for s in sections)
    logger.info(
        "Wrote Excel report -> %s (%d sheet(s), %d row(s) total)",
        output_path, len(sections), total_rows,
    )
    return output_path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _sanitize_sheet_name(raw: str) -> str:
    """Make a sheet name safe and within Excel's 31-char limit."""
    forbidden = set(r"[]:*?/\\")
    cleaned = "".join("_" if ch in forbidden else ch for ch in raw)
    return cleaned[:31] or "Sheet"


def _write_headers(ws: Worksheet, headers: list[str]) -> None:
    """Render the styled header row."""

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _write_body(
    ws: Worksheet,
    headers: list[str],
    body: list[list[str]],
) -> None:
    """Stream the body rows beneath the header.

    Numeric-looking cells are right-aligned for readability. We classify
    a column as numeric by inspecting the *header* keywords - this is
    cheap and works for our deterministic schemas.
    """

    numeric_columns = {
        idx for idx, header in enumerate(headers, start=1)
        if any(token in header.lower() for token in (
            "price", "amount", "shares", "cost", "proceeds",
                    "gain", "loss", "invested", "%",
        ))
    }

    for row_idx, row in enumerate(body, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in numeric_columns:
                cell.alignment = _NUMERIC_ALIGNMENT


def _apply_column_widths(
    ws: Worksheet,
    headers: list[str],
    body: list[list[str]],
) -> None:
    """Best-effort auto-sizing using the longest content per column.

    openpyxl has no real auto-fit so we approximate it. A small padding
    factor avoids cramped column widths.
    """

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in body:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(row[col_idx - 1]))
        # Cap the width so very long ISIN strings do not blow up the sheet.
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max_len + 2, 40
        )
